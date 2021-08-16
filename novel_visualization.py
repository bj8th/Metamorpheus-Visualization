#%%
import os
from pyteomics import mzml
import logging
import spectra_plot
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import Workbook

def filter_mzml(mzml_filepath, novel_scan_numbers):
    filtered_mzml = {}
    for mzml_spectra in mzml.MzML(mzml_filepath):
        scan_number = int(mzml_spectra['id'].split('scan=')[1])
        if scan_number in novel_scan_numbers:
            filtered_mzml[scan_number] = mzml_spectra
    return filtered_mzml

def extract_mzml_data(spectra):
    spectra_dict = {}
    try:
        spectra_dict['precursor charge state'] = int(spectra['precursorList']['precursor'][0]['selectedIonList']['selectedIon'][0]['charge state'])
        spectra_dict['precursor m/z'] = float(spectra['precursorList']['precursor'][0]['selectedIonList']['selectedIon'][0]['selected ion m/z'])
    except:
        logging.warn(f"Precuror information not found for {peptide}")
    spectra_dict['lowest observed m/z'] = spectra['lowest observed m/z']
    spectra_dict['highest observed m/z'] = spectra['highest observed m/z']
    spectra_dict['m/z array'] = spectra['m/z array']
    spectra_dict['intensity array'] = spectra['intensity array']
    spectra_dict['base peak m/z'] = spectra['base peak m/z']
    spectra_dict['minimum intensity'] = min(spectra['intensity array'])
    return spectra_dict

def separate_matching_ions(ions, type_convert=float):
    ion_type_list = ions.split(';')
    matched_ions = {}
    for ion_type in ion_type_list:
        ion_map = ion_type.strip('][').split(', ')
        
        for ion in ion_map:
            tmp = ion.split(':')
            matched_ions[tmp[0]] = type_convert(tmp[1])
    return matched_ions

def combine_matches(row):
    combined = {}
    combined = {key:{'m/z':mzml, 'intensity':row['intensity_match'][key]} for key, mzml in row['mz_match'].items()}
    # for key in mzml.keys():
    #     print(key)
    #     print(mzml)
    #     return mzml
    #     combined[key]['m/z'] = mzml[key]
    #     combined[key]['intensity'] = intensity[key]
    return combined
def process_metamorpheus_psm(psm_filepath, novel_scans):
    psm = pd.read_table(psm_filepath)
    filtered_psms = []
    for scan in novel_scans:
        tmp = psm
        for column, value in scan.items():
            tmp = tmp[tmp[column] == value]
        filtered_psms.append(tmp)
    psm = pd.concat(filtered_psms)
    # psm = psm[psm['Scan Number'].isin(novel_peptides)]
    psm['mz_match'] = psm['Matched Ion Mass-To-Charge Ratios'].apply(separate_matching_ions)
    psm['intensity_match'] = psm['Matched Ion Intensities'].apply(separate_matching_ions)
    psm['matched_ions'] = psm.apply(combine_matches, axis = 1)
    return psm 


def plot_mirror_into_exel(novel_scans, psm, mzml_directory):
    if not os.path.exists('individual_mirror_plots'):
        os.mkdir('individual_mirror_plots')
    filename_column = 'A'
    scan_number_column = 'B'
    peptide_base_column = 'C'
    peptide_full_column = 'D'
    accession_column = 'E'
    gene_column = 'F'
    # score_column = 'G'
    # qvalue_column = 'H'
    mirrorplot_column='G'
    
    wb = Workbook()
    ws = wb.active

    ws.append([
        "Filename", 
        "Scan Number", 
        "Base Sequence",
        "Full Sequence", 
        "Accession",
        "Gene", 
        # "Score", 
        # "Q-Value", 
        "Mirror Plot"])
    ws.row_dimensions[2].height=375
    ws.column_dimensions[mirrorplot_column].width = 80
    excel_row_number = 2
    for filename, group in psm.groupby('File Name'):
        scans = list(group['Scan Number'])

        mzml_filepath=os.path.join(mzml_directory, filename) + '.mzML'
        novel_mzml = filter_mzml(mzml_filepath, scans)
        for index, row in group.iterrows():
            scan = row['Scan Number']
            mzml_data = extract_mzml_data(novel_mzml[scan])
            mirror = spectra_plot.SpectraPlot(
                mass_charge_ratios=mzml_data['m/z array'],
                intensities = mzml_data['intensity array'])
            mirror.labelled_spectra = row['matched_ions']
            mirror.filter()
            ax = mirror.mirror_plot(label=True,adjust_annotation = False)
            plt.savefig(f'individual_mirror_plots/mass-spec_{scan}.png')
            img = Image(f'individual_mirror_plots/mass-spec_{scan}.png')
            ws.add_image(img, f'{mirrorplot_column}{excel_row_number}')
            ws.row_dimensions[excel_row_number].height=375

            ws[f'{filename_column}{excel_row_number}']=row['File Name']
            ws[f'{scan_number_column}{excel_row_number}']=row['Scan Number']
            ws[f'{peptide_base_column}{excel_row_number}']=row['Base Sequence']
            ws[f'{peptide_full_column}{excel_row_number}']=row['Full Sequence']
            ws[f'{accession_column}{excel_row_number}']=row['Protein Accession']
            ws[f'{gene_column}{excel_row_number}']=row['Gene Name']
            excel_row_number = excel_row_number + 1
    wb.save(f'novel_peptides.xlsx')

#%%
psm_filepath = '/Users/bj8th/Documents/Sheynkman-Lab/GitHub/Long-Read-Proteogenomics/results/test_jurkat_chr22_with_sqanti/metamorpheus/pacbio/hybrid/search_results/Task1SearchTask/AllPSMs.tsv'
mzml_filepath = '/Users/bj8th/Documents/Sheynkman-Lab/Data/test/MassSpec/mzml/120426_Jurkat_highLC_Frac22.mzML'
novel_filepath = ''

novel_scans = pd.DataFrame({
    'scan_num' : [17192, 17342, 15689], 
    'file_name':['120426_Jurkat_highLC_Frac22','120426_Jurkat_highLC_Frac22','120426_Jurkat_highLC_Frac22'], 
    'gene': ['gene1','gene2','gene3'],
    'acc':['PB.1.1', 'PB.1.2', 'PB.1.3'],
    'seq':['YABBA','DABBA','DO'], 
    'score':[0,0,0], 
    'qval':[0,0,0], 
    'gene_name':['gene1','gene2','gene3'], 
    'ec_priority':[0,0,0]
    })

novel_scans_list = []
for index, row in novel_scans.iterrows():
    novel_scans_list.append({'File Name' : row['file_name'], 'Scan Number': row['scan_num']})
psm = process_metamorpheus_psm(psm_filepath, novel_scans_list)

mzml_dir = '/Users/bj8th/Documents/Sheynkman-Lab/Data/test/MassSpec/mzml/'
plot_mirror_into_exel(novel_scans, psm, mzml_dir)
